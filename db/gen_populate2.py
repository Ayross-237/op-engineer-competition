"""Generate db/populate2.sql from the data in resources/.

Reads resources/{caterers,sessions,students}.xlsx and bakes in the contact /
menu / absence / exclusion data that lives in the PDFs. Outputs SQL that
matches the current schema (db/schema.sql).

Re-run after editing the resources or the schema:
    python db/gen_populate2.py > db/populate2.sql
"""
import random
import sys
from datetime import datetime, time, timedelta
from pathlib import Path

import openpyxl

RESOURCES = Path(__file__).resolve().parents[1] / "resources"


# --- Static data from PDFs (caterer-contacts, caterer-menus) ---

CATERERS = [
    # (name, region, contact_email, chef_email, cc_chef)
    ("Lakehouse Victoria Point", "Redlands",         "carmen@padea.com.au",        None,                       False),
    ("Terrific Noodles",         "South Brisbane",   "cherndylan@gmail.com",       "dylanchern808@gmail.com",  False),
    ("Kenko Sushi House",        "West Brisbane",    "hellopadea@gmail.com",       None,                       False),
    ("Guzman y Gomez",           "Central Brisbane", "carmengabrielleee@gmail.com","dylan@padea.com.au",       True),
]

# Pricing: (caterer_name, price_per_item, per_trip_fee, per_school_per_trip_fee)
# Lakehouse: $35 ex GST per item, $0 delivery.
# Terrific Noodles: $20.50 ex GST per item, $30 per school per trip.
# Kenko: $5.50 inc GST per item, $10 per school per trip.
# Guzman y Gomez: $15 inc GST per item, $50 per trip (flat, not per school).
PRICING = [
    ("Lakehouse Victoria Point", 35.00, 0.00, 0.00),
    ("Terrific Noodles",         20.50, 0.00, 30.00),
    ("Kenko Sushi House",         5.50, 0.00, 10.00),
    ("Guzman y Gomez",           15.00, 50.00, 0.00),
]

# Menus. PDF tags: GF, DF, NF, VO (= vegetarian option → V in our enum).
# Per the PDF: "Assume all non-pork meals are halal" — so any dish whose name
# doesn't obviously contain pork/bacon gets H added.
MENUS = {
    "Lakehouse Victoria Point": [
        ("Shrimp Fried Rice",                "GF DF"),
        ("Spaghetti Bolognese + Garlic Bread","NF"),
        ("Sweet and Sour Chicken",           "GF DF NF"),
        ("Classic Cream Pasta",              "NF"),
        ("Gnocchi in Tomato Sauce",          "NF"),
        ("Chicken, Bacon, Avo Wrap",         "VO"),       # bacon → not halal
        ("Fried Chicken Burger + Chips",     "NF"),
        ("Fish Taco Bowl",                   "NF"),
        ("Korean Beef Bulgogi Rice Bowl",    "GF DF NF"),
        ("Japanese Chicken Curry",           "DF NF VO"),
    ],
    "Terrific Noodles": [
        ("Spicy Miso Udon",                  "DF"),
        ("Stir-fry Noodles topped with Chicken","GF DF NF"),
        ("Grilled Pork Vermicelli Salad",    "GF DF NF VO"),  # pork → not halal
        ("Spaghetti meatballs",              "NF"),
        ("Lemongrass Grilled Beef and Noodles","GF DF NF VO"),
        ("Creamy Garlic Beef Noodles",       "VO"),
        ("Mie Goreng",                       "GF DF NF VO"),
        ("Beef Pad Thai",                    ""),
        ("Bacon Carbonara",                  ""),               # bacon → not halal
        ("Chinese Honey Soy Noodles",        "DF"),
    ],
    "Kenko Sushi House": [
        ("Lamb wrap",                        "NF"),
        ("Chicken Parmi, chips and salad",   "DF NF"),
        ("Japanese Chicken Katsu",           "NF VO"),
        ("Teriyaki Salmon rice bowl",        "GF DF NF VO"),
        ("Chicken Karaage ricebowl",         "DF NF VO"),
        ("Creamy Udon",                      ""),
        ("Beef Fried Rice",                  "GF DF VO"),
        ("Mongolian Beef and Rice",          ""),
        ("Sweet and Sour Chicken",           "GF DF"),
        ("Chinese Honey Soy Noodles",        "DF"),
    ],
    "Guzman y Gomez": [
        ("Breakfast Tacos",                  "NF VO"),
        ("Caesar Salad",                     "GF DF NF VO"),
        ("Cali Burrito",                     ""),
        ("Grilled Chicken Burrito",          ""),
        ("Pulled pork burrito bowl",         "GF NF VO"),   # pork → not halal (non-veg version)
        ("Nachos",                           "GF VO"),
        ("Nacho Fries",                      "GF VO"),
        ("Chicken Quesadilla",               ""),
        ("Chicken Enchilada",                "GF DF"),
        ("Crispy Chicken Taco",              ""),
    ],
}

# Dishes that contain pork/bacon and therefore are NOT halal.
NON_HALAL = {
    "Chicken, Bacon, Avo Wrap",
    "Grilled Pork Vermicelli Salad",
    "Bacon Carbonara",
    "Pulled pork burrito bowl",
}

# School → caterer mapping (from sessions.xlsx, deduped).
SCHOOL_CATERER = {
    "Moreton Bay Boys' College":      "Lakehouse Victoria Point",
    "John Paul College":              "Terrific Noodles",
    "MacGregor State High School":    "Terrific Noodles",
    "Indooroopilly State High School":"Kenko Sushi House",
    "Loreto College":                 "Guzman y Gomez",
    "Cannon Hill Anglican College":   "Guzman y Gomez",
}

# Absences from absences.pdf: (school_name, date_iso, student_name)
ABSENCES = [
    ("Moreton Bay Boys' College",      "2026-05-02", "Noah Baker"),
    ("John Paul College",              "2026-05-02", "Christina Hu"),
    ("John Paul College",              "2026-05-02", "Nathan Smith"),
    ("MacGregor State High School",    "2026-05-04", "Rose Smith"),
    ("Indooroopilly State High School","2026-05-02", "Charlie Morris"),
    ("Indooroopilly State High School","2026-05-02", "Jack Carter"),
    ("Indooroopilly State High School","2026-05-02", "Charlie Mitchell"),
    ("Loreto College",                 "2026-05-01", "Holly Hill"),
    ("Loreto College",                 "2026-05-01", "Imogen Evans"),
    ("Cannon Hill Anglican College",   "2026-05-03", "Henry Cook"),
]

# Exclusions from exclusions.pdf.
# (school, date_iso, cancelled_year_levels|None) — None means whole-session cancellation.
EXCLUSIONS = [
    ("Indooroopilly State High School","2026-05-04", None),       # all year levels
    ("Loreto College",                 "2026-05-02", None),       # all year levels
    ("Cannon Hill Anglican College",   "2026-05-03", [12, 10]),   # Year 11 still attends
]

# Fake feedback (manager-submitted notes about each caterer). 10 per caterer,
# spanning ~10 weeks before the May 2026 sessions. Each entry mentions specific
# dishes from that caterer's menu so an LLM can derive per-dish ratings.
FEEDBACK = [
    # --- Lakehouse Victoria Point (Tuesdays) ---
    ("Lakehouse Victoria Point", "2026-02-24 19:30+10", "Shrimp Fried Rice was the standout tonight — generous portions, prawns were fresh and rice was fluffy. Spaghetti Bolognese arrived a bit cold but the garlic bread was crisp. Chicken, Bacon, Avo Wrap was disappointing — wraps had gone soggy in transit and the lettuce had wilted. Japanese Chicken Curry was a hit with Year 12s, sauce had real depth. Korean Beef Bulgogi was tender and well-marinated. Strong overall night."),
    ("Lakehouse Victoria Point", "2026-03-03 19:30+10", "Fried Chicken Burger + Chips was outstanding — crispy chicken, hot chips that stayed crunchy. Classic Cream Pasta was the weakest item, sauce had broken and pasta clumped together. Sweet and Sour Chicken was decent but the sauce was a touch too sweet. Korean Beef Bulgogi continues to be reliable. Gnocchi in Tomato Sauce was undercooked — gnocchi had a chalky centre."),
    ("Lakehouse Victoria Point", "2026-03-10 19:30+10", "Excellent night for the curries — Japanese Chicken Curry was the best version yet, sauce rich and chicken tender. Shrimp Fried Rice was solid. Fish Taco Bowl was bright and zesty, students loved it. Chicken, Bacon, Avo Wrap was better than past weeks, wraps held together. Spaghetti Bolognese was over-portioned and several students couldn't finish."),
    ("Lakehouse Victoria Point", "2026-03-17 19:30+10", "Off night. Korean Beef Bulgogi Rice Bowl was tough — beef overcooked across most boxes. Fried Chicken Burger + Chips was the saving grace. Classic Cream Pasta was bland, lacking salt. Japanese Chicken Curry vegetarian version was watery and under-seasoned. Fish Taco Bowl was good but portions were small. Lakehouse should review their beef cooking time."),
    ("Lakehouse Victoria Point", "2026-03-24 19:30+10", "Sweet and Sour Chicken sauce was perfect tonight — well-balanced tang. Shrimp Fried Rice was great. Gnocchi in Tomato Sauce finally cooked through properly, much better than last week. Spaghetti Bolognese was solid. Chicken, Bacon, Avo Wrap was lukewarm by service time. Overall a strong recovery from last week."),
    ("Lakehouse Victoria Point", "2026-03-31 19:30+10", "Japanese Chicken Curry was excellent both versions — rich, well-spiced, the vegetarian had a good aubergine swap. Korean Beef Bulgogi was tender and beautifully marinated. Fried Chicken Burger + Chips was great, chips stayed crispy. Fish Taco Bowl was the lowest pick of the night, only two ordered. Classic Cream Pasta was good but unremarkable."),
    ("Lakehouse Victoria Point", "2026-04-07 19:30+10", "Shrimp Fried Rice was lukewarm — heat retention is slipping again. Fried Chicken Burger was solid. Gnocchi in Tomato Sauce had a beautiful sauce but the gnocchi was sticky. Korean Beef Bulgogi was perfect — best of the term so far, beef practically melted. Chicken, Bacon, Avo Wrap was fine, no complaints."),
    ("Lakehouse Victoria Point", "2026-04-14 19:30+10", "Spaghetti Bolognese was outstanding tonight — meat sauce was rich, garlic bread perfectly golden. Sweet and Sour Chicken was great, well-balanced sauce. Japanese Chicken Curry was reliable. Fish Taco Bowl came with broken taco shells, presentation poor though flavour was fine. Classic Cream Pasta was the worst item again, sauce had separated."),
    ("Lakehouse Victoria Point", "2026-04-21 19:30+10", "Excellent across the board. Korean Beef Bulgogi was the standout — best execution yet, students consistently picked it first. Japanese Chicken Curry was great. Fried Chicken Burger was crispy and hot. Sweet and Sour Chicken was solid. Gnocchi in Tomato Sauce was perfect tonight. Shrimp Fried Rice was good."),
    ("Lakehouse Victoria Point", "2026-04-28 19:30+10", "Strong end to the term. Japanese Chicken Curry was perfect — both versions. Korean Beef Bulgogi was tender. Fried Chicken Burger + Chips was great. Spaghetti Bolognese was rich and well-portioned. Fish Taco Bowl was the only weak point — fish was overcooked and dry. Classic Cream Pasta still inconsistent — needs a recipe review."),

    # --- Terrific Noodles (Tuesdays / Wednesdays) ---
    ("Terrific Noodles", "2026-02-24 20:00+10", "Spicy Miso Udon was excellent tonight — broth had real depth and noodles had bite. Mie Goreng was solid both versions. Beef Pad Thai was great. Bacon Carbonara was bland and the sauce was watery. Grilled Pork Vermicelli Salad was fresh, herbs at peak. Lemongrass Grilled Beef and Noodles was the standout — perfectly grilled beef with bright dressing."),
    ("Terrific Noodles", "2026-03-03 20:00+10", "Off night for Beef Pad Thai — noodles clumped, sauce pooled at the bottom of the container. Spicy Miso Udon broth was diluted, watery. Stir-fry Noodles topped with Chicken was on point, well-seasoned. Mie Goreng was excellent. Creamy Garlic Beef Noodles was rich and indulgent. Chinese Honey Soy Noodles was solid. Late delivery cost us 10 minutes of dinner break."),
    ("Terrific Noodles", "2026-03-10 20:00+10", "Mie Goreng was outstanding — wok-hei was present, noodles had character. Beef Pad Thai recovered from last week. Lemongrass Grilled Beef and Noodles was great. Bacon Carbonara was finally well-executed, sauce coated noodles properly. Spaghetti meatballs were overcooked and mushy. Chinese Honey Soy Noodles was reliable."),
    ("Terrific Noodles", "2026-03-17 20:00+10", "Strong night across the board. Spicy Miso Udon was excellent. Stir-fry Noodles topped with Chicken was crisp and well-seasoned. Lemongrass Grilled Beef both versions worked well. Grilled Pork Vermicelli Salad was fresh. Creamy Garlic Beef Noodles was rich. Chinese Honey Soy Noodles was the weakest — too salty tonight."),
    ("Terrific Noodles", "2026-03-24 20:00+10", "Beef Pad Thai was the standout, possibly best of the term — moist, balanced, peanuts crunchy. Mie Goreng was great both versions. Spicy Miso Udon broth was thicker than usual, students loved it. Bacon Carbonara was inconsistent — some servings creamy, others dry. Spaghetti meatballs were tender but the sauce lacked depth."),
    ("Terrific Noodles", "2026-03-31 20:00+10", "Stir-fry Noodles topped with Chicken was excellent. Lemongrass Grilled Beef and Noodles vegetarian was the best vegetarian version so far — marinated mushrooms held flavour. Mie Goreng was good. Spicy Miso Udon was reliable. Creamy Garlic Beef Noodles was over-rich tonight, students struggled to finish. Chinese Honey Soy Noodles was great."),
    ("Terrific Noodles", "2026-04-07 20:00+10", "Spicy Miso Udon was disappointing tonight — broth was thin and lacked the usual depth. Beef Pad Thai was outstanding — moist, balanced, beef tender. Mie Goreng both versions were strong. Spaghetti meatballs improved over recent weeks. Lemongrass Grilled Beef and Noodles was perfect. Grilled Pork Vermicelli Salad was a fresh contrast on a heavy menu."),
    ("Terrific Noodles", "2026-04-14 20:00+10", "Strong night. Mie Goreng was excellent. Stir-fry Noodles topped with Chicken was on form. Beef Pad Thai was good. Bacon Carbonara was creamy and rich, best version of the term. Grilled Pork Vermicelli Salad was fresh and bright. Chinese Honey Soy Noodles continues to be a reliable middle-pack option."),
    ("Terrific Noodles", "2026-04-21 20:00+10", "Outstanding night for Lemongrass Grilled Beef and Noodles — best execution of the term, both versions. Spicy Miso Udon was great. Beef Pad Thai was solid. Mie Goreng was reliable. Creamy Garlic Beef Noodles was perfect tonight, sauce was velvety. Spaghetti meatballs were lukewarm — heat retention issue persists."),
    ("Terrific Noodles", "2026-04-28 20:00+10", "Closing strong. Mie Goreng was outstanding both versions. Spicy Miso Udon was at peak. Beef Pad Thai was great. Lemongrass Grilled Beef and Noodles non-veg was perfect; veg version slightly under-seasoned. Stir-fry Noodles topped with Chicken was good. Bacon Carbonara was the weakest — sauce broke again."),

    # --- Kenko Sushi House (Mondays) ---
    ("Kenko Sushi House", "2026-02-23 19:00+10", "Teriyaki Salmon rice bowl was outstanding — salmon was buttery, glaze deep. Chicken Karaage ricebowl was crispy. Sweet and Sour Chicken was reliable. Lamb wrap was excellent — flavoursome, well-spiced. Beef Fried Rice was good. Mongolian Beef and Rice was tough, beef overcooked. Best vegetarian variety we've seen from a caterer in weeks."),
    ("Kenko Sushi House", "2026-03-02 19:00+10", "Chicken Parmi, chips and salad was the standout — schnitzel was crispy, cheese melted properly, chips stayed warm. Teriyaki Salmon rice bowl was great. Japanese Chicken Katsu was solid. Creamy Udon was bland — broth lacked depth. Lamb wrap was perfect. Chicken Karaage ricebowl arrived less crispy than usual, packaging let it down."),
    ("Kenko Sushi House", "2026-03-09 19:00+10", "Strong night. Teriyaki Salmon rice bowl was excellent. Chicken Karaage ricebowl was crispy and the dipping sauce had kick. Beef Fried Rice was good. Mongolian Beef and Rice was much better tonight, beef was tender after longer braising. Japanese Chicken Katsu was reliable. Sweet and Sour Chicken was lukewarm by service."),
    ("Kenko Sushi House", "2026-03-16 19:00+10", "Off night. Salmon in the Teriyaki bowl was overcooked and dry. Chicken Parmi, chips and salad was the saving grace. Lamb wrap was solid. Beef Fried Rice was good. Mongolian Beef and Rice was tough again — needs longer braising. Chinese Honey Soy Noodles was decent. Heat retention generally good across boxes."),
    ("Kenko Sushi House", "2026-03-23 19:00+10", "Excellent across the board. Teriyaki Salmon rice bowl was perfect — back on form. Chicken Karaage ricebowl was crispy and juicy. Lamb wrap was outstanding — best of the term. Beef Fried Rice was great. Mongolian Beef and Rice was tender finally. Best night of the term for Kenko."),
    ("Kenko Sushi House", "2026-03-30 19:00+10", "Chicken Parmi, chips and salad continues to be reliable. Teriyaki Salmon rice bowl was good. Japanese Chicken Katsu was great — perfectly fried and well-portioned. Lamb wrap was solid. Sweet and Sour Chicken was over-sauced tonight. Creamy Udon was thicker than usual and students appreciated the change."),
    ("Kenko Sushi House", "2026-04-06 19:00+10", "Teriyaki Salmon rice bowl was outstanding. Chicken Karaage ricebowl was crispy. Lamb wrap was excellent. Beef Fried Rice was good. Mongolian Beef and Rice was tender. Best vegetarian execution of any caterer this week — both veg versions matched non-veg quality, the marinated tofu element in the Teriyaki worked particularly well."),
    ("Kenko Sushi House", "2026-04-13 19:00+10", "Quiet night. Teriyaki Salmon was great. Chicken Karaage ricebowl was solid. Chicken Parmi, chips and salad was the standout — perfectly cooked, generously portioned. Lamb wrap was good. Sweet and Sour Chicken was excellent tonight, sauce well-balanced. Creamy Udon broth was thin again — recurring weakness."),
    ("Kenko Sushi House", "2026-04-20 19:00+10", "Outstanding night. Lamb wrap was best of the term. Teriyaki Salmon rice bowl was buttery. Chicken Karaage was the crispiest yet. Beef Fried Rice was great. Mongolian Beef and Rice was tender and well-flavoured. Heat retention perfect across all boxes. Kenko continues to set the standard for the price point."),
    ("Kenko Sushi House", "2026-04-27 19:00+10", "Strong close. Teriyaki Salmon rice bowl was excellent both versions. Chicken Parmi was crispy. Chicken Karaage ricebowl was great. Lamb wrap was reliable. Beef Fried Rice was good. Mongolian Beef and Rice was perfect — tender, well-marinated. Sweet and Sour Chicken was the only weak point — sauce ran thin."),

    # --- Guzman y Gomez (Mondays) ---
    ("Guzman y Gomez", "2026-02-23 19:00+10", "Pulled pork burrito bowl was the standout — pork was tender, beans well-seasoned. Nachos were soggy by service time, chips had lost crunch in transit. Caesar Salad was crisp, dressing well-balanced. Chicken Enchilada was hot and cheesy. Cali Burrito was good but a touch dry. Crispy Chicken Taco was excellent."),
    ("Guzman y Gomez", "2026-03-02 19:00+10", "Caesar Salad was excellent. Chicken Quesadilla was hot and cheesy. Pulled pork burrito bowl was great. Breakfast Tacos were the weakest — egg was overcooked. Nachos arrived soggy again — packaging needs work, recommend separating sour cream and salsa. Grilled Chicken Burrito was solid. Crispy Chicken Taco was crunchy and well-seasoned."),
    ("Guzman y Gomez", "2026-03-09 19:00+10", "Strong night. Pulled pork burrito bowl was outstanding. Chicken Enchilada was perfect — cheese stretched beautifully. Nachos finally arrived in deconstructed packaging, chips stayed crisp. Caesar Salad was great. Crispy Chicken Taco was excellent. Nacho Fries were the surprise hit — crispy and well-topped."),
    ("Guzman y Gomez", "2026-03-16 19:00+10", "Cali Burrito was excellent tonight — rice was fluffy, beans well-seasoned. Chicken Quesadilla was crispy and cheesy. Grilled Chicken Burrito was solid. Caesar Salad was crisp. Pulled pork burrito bowl was good. Nachos slipped back to soggy assembly — packaging consistency issue persists."),
    ("Guzman y Gomez", "2026-03-23 19:00+10", "Off night. Caesar Salad arrived limp, dressing applied too early and the romaine wilted. Chicken Enchilada was over-sauced, sauce pooled at the bottom. Nachos were soggy. Pulled pork burrito bowl was the saving grace. Breakfast Tacos were better than last week. Crispy Chicken Taco was good. Nacho Fries were excellent."),
    ("Guzman y Gomez", "2026-03-30 19:00+10", "Recovery night. Chicken Enchilada was excellent — best version yet. Pulled pork burrito bowl was outstanding both versions. Nachos correctly assembled, crispy. Caesar Salad was perfect. Cali Burrito was good. Crispy Chicken Taco was crunchy. Grilled Chicken Burrito was solid."),
    ("Guzman y Gomez", "2026-04-06 19:00+10", "Pulled pork burrito bowl was great. Chicken Quesadilla was excellent. Caesar Salad was crisp and well-dressed both versions. Nachos were good. Nacho Fries were the standout — crispy, well-topped. Chicken Enchilada was solid. Crispy Chicken Taco continues to be a reliable favourite."),
    ("Guzman y Gomez", "2026-04-13 19:00+10", "Outstanding night. Chicken Enchilada was perfect. Pulled pork burrito bowl was excellent. Nachos correctly assembled and well-topped. Caesar Salad was great. Crispy Chicken Taco was crunchy. Breakfast Tacos were finally well-executed — eggs were soft, tortillas warm. Cali Burrito was solid."),
    ("Guzman y Gomez", "2026-04-20 19:00+10", "Strong night. Pulled pork burrito bowl outstanding both versions. Caesar Salad was perfect. Chicken Quesadilla was hot and cheesy. Nacho Fries were great. Chicken Enchilada was excellent. Grilled Chicken Burrito was solid. Nachos slipped back to soggy — packaging inconsistency is the recurring complaint."),
    ("Guzman y Gomez", "2026-04-27 19:00+10", "Closing strong. Chicken Enchilada was perfect — cheese-pull worthy. Pulled pork burrito bowl was great both versions. Caesar Salad was crisp. Nachos finally arrived crispy. Crispy Chicken Taco was excellent. Breakfast Tacos were solid. Nacho Fries were the standout — crispy and well-portioned. GYG closes the term on a high."),
]

# Sheet name → (school name, day label) used to map student sheets to programs.
SHEET_TO_PROGRAM = {
    "MBBC":             ("Moreton Bay Boys' College",      "Tuesday"),
    "JPC - Tuesday":    ("John Paul College",              "Tuesday"),
    "JPC - Wednesday":  ("John Paul College",              "Wednesday"),
    "MSHS":             ("MacGregor State High School",    "Thursday"),
    "ISHS - Monday":    ("Indooroopilly State High School","Monday"),
    "ISHS - Tuesday":   ("Indooroopilly State High School","Tuesday"),
    "ISHS - Thursday":  ("Indooroopilly State High School","Thursday"),
    "LC - Monday":      ("Loreto College",                 "Monday"),
    "LC - Tuesday":     ("Loreto College",                 "Tuesday"),
    "CHAC - Monday":    ("Cannon Hill Anglican College",   "Monday"),
    "CHAC - Wednesday": ("Cannon Hill Anglican College",   "Wednesday"),
}


# --- helpers ---

# The FEEDBACK prose was authored against Feb–Apr 2026 sessions. Shift the whole series
# forward by this many whole weeks so the latest entries land in early June 2026 — recent
# enough to populate the caterer-facing "last N weeks" feedback summary. Whole weeks keep
# each entry's weekday, which dish-rating generation relies on to map a feedback date to
# the programs running that day.
FEEDBACK_SHIFT_WEEKS = 5

def shift_submitted_at(submitted_at, weeks):
    """Shift a feedback timestamp forward by whole weeks, preserving the weekday and the
    original time/timezone suffix (e.g. '2026-02-24 19:30+10' -> '2026-03-31 19:30+10')."""
    date_part, sep, rest = submitted_at.partition(" ")
    shifted = datetime.strptime(date_part, "%Y-%m-%d").date() + timedelta(weeks=weeks)
    return f"{shifted.isoformat()}{sep}{rest}"

def sql_str(s):
    """Escape a value for inclusion in SQL. None → NULL."""
    if s is None:
        return "NULL"
    if isinstance(s, bool):
        return "TRUE" if s else "FALSE"
    if isinstance(s, (int, float)):
        return str(s)
    return "'" + str(s).replace("'", "''") + "'"


def sql_array(values):
    """Render a list of strings as a Postgres text array literal: '{A,B,C}'."""
    return "'{" + ",".join(values) + "}'"


def parse_time(s):
    """Parse '4:30pm' or '3:15pm' etc. into a 24h HH:MM string."""
    s = s.strip().lower()
    if s.endswith("am") or s.endswith("pm"):
        meridiem = s[-2:]
        rest = s[:-2].strip()
        if ":" in rest:
            h, m = rest.split(":")
            h, m = int(h), int(m)
        else:
            h, m = int(rest), 0
        if meridiem == "pm" and h != 12:
            h += 12
        if meridiem == "am" and h == 12:
            h = 0
        return f"{h:02d}:{m:02d}"
    return s


def parse_dietary(text):
    """Parse a dietary cell into (tags, extra, wants_catering).

    Accepts a comma-separated mix of enum labels and free text. Enum hits
    populate tags[]; 'Opted out of Catering' flips wants_catering; anything
    else goes into dietary_extra.
    """
    if text is None or not str(text).strip():
        return [], None, True
    tags = []
    extras = []
    wants = True
    for raw in str(text).split(","):
        p = raw.strip()
        lo = p.lower()
        if lo == "gluten free":
            tags.append("GF")
        elif lo == "dairy free":
            tags.append("DF")
        elif lo == "nut free":
            tags.append("NF")
        elif lo == "vegetarian":
            tags.append("V")
        elif lo == "halal":
            tags.append("H")
        elif lo == "opted out of catering":
            wants = False
        else:
            extras.append(p)
    return tags, ", ".join(extras) if extras else None, wants


def menu_tags(pdf_tags, dish_name):
    """Convert PDF tag string like 'GF DF VO' into our enum array, applying
    the halal-by-default rule (any non-pork/non-bacon dish gets H)."""
    out = []
    for t in pdf_tags.split():
        if t == "VO":
            out.append("V")
        elif t in {"GF", "DF", "NF"}:
            out.append(t)
    if dish_name not in NON_HALAL:
        out.append("H")
    return out


# --- main generation ---

def main():
    out = []
    p = out.append

    p("-- Generated by db/gen_populate2.py from resources/{caterers,sessions,students}.xlsx")
    p("-- plus the data baked in to that script from caterer-contacts.pdf, caterer-menus.pdf,")
    p("-- exclusions.pdf and absences.pdf. Do not edit by hand — re-run the generator.")
    p("-- Assumes a clean DB (run db/reset.sql then db/schema.sql first).")
    p("")

    # --- caterers ---
    p("-- Caterers → IDs 1..4")
    p("INSERT INTO caterers (name, region, contact_email, chef_email, cc_chef) VALUES")
    rows = []
    for name, region, contact, chef, cc in CATERERS:
        rows.append(f"    ({sql_str(name)}, {sql_str(region)}, {sql_str(contact)}, {sql_str(chef)}, {sql_str(cc)})")
    p(",\n".join(rows) + ";")
    p("")

    caterer_id = {name: i + 1 for i, (name, *_) in enumerate(CATERERS)}

    # --- pricing ---
    p("-- Pricing structures (1:1 with caterers)")
    p("INSERT INTO pricing_structures (caterer_id, price_per_item, per_trip_fee, per_school_per_trip_fee) VALUES")
    rows = []
    for name, ppi, ptf, pspt in PRICING:
        rows.append(f"    ({caterer_id[name]}, {ppi:.2f}, {ptf:.2f}, {pspt:.2f})")
    p(",\n".join(rows) + ";")
    p("")

    # --- items ---
    p("-- Menu items per caterer. H added to every dish whose name does not contain")
    p("-- 'pork' or 'bacon' (assume non-pork = halal). A source VO (vegetarian option)")
    p("-- dish is split into two orderable items: the default version (no V tag) and a")
    p("-- '... (vegetarian)' version that carries the V tag, since VO means the caterer")
    p("-- can prepare it vegetarian on request rather than it being vegetarian by default.")
    p("INSERT INTO items (caterer_id, name, dietary_tags) VALUES")
    rows = []
    caterer_items = {}  # caterer name -> list of (item_name, tags) as actually inserted
    for cname, dishes in MENUS.items():
        cid = caterer_id[cname]
        items_for_caterer = []
        for dish_name, pdf_tags in dishes:
            tags = menu_tags(pdf_tags, dish_name)
            if "V" in tags:
                default_tags = [t for t in tags if t != "V"]
                rows.append(f"    ({cid}, {sql_str(dish_name)}, {sql_array(default_tags)})")
                rows.append(f"    ({cid}, {sql_str(dish_name + ' (vegetarian)')}, {sql_array(tags)})")
                items_for_caterer.append((dish_name, default_tags))
                items_for_caterer.append((dish_name + " (vegetarian)", tags))
            else:
                rows.append(f"    ({cid}, {sql_str(dish_name)}, {sql_array(tags)})")
                items_for_caterer.append((dish_name, tags))
        caterer_items[cname] = items_for_caterer
    p(",\n".join(rows) + ";")
    p("")

    # --- schools ---
    p("-- Schools. caterer_id assigned per sessions.xlsx (each school currently uses one caterer).")
    p("INSERT INTO schools (name, region, caterer_id) VALUES")
    school_id = {}
    sessions_wb = openpyxl.load_workbook(RESOURCES / "sessions.xlsx", data_only=True)
    sessions_ws = sessions_wb["sessions"]
    school_region = {}
    for row in sessions_ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        school_region[row[0]] = row[1]
    rows = []
    school_id_counter = 1
    for sname in SCHOOL_CATERER:
        school_id[sname] = school_id_counter
        school_id_counter += 1
        region = school_region.get(sname, "")
        rows.append(f"    ({sql_str(sname)}, {sql_str(region)}, {caterer_id[SCHOOL_CATERER[sname]]})")
    p(",\n".join(rows) + ";")
    p("")

    # --- programs ---
    # Group sessions.xlsx rows by (school, day) to derive programs.
    p("-- Programs (weekly recurring slots). Derived from sessions.xlsx by grouping rows")
    p("-- on (school, day-of-week). Each program inherits manager/time/year-levels from")
    p("-- one of its rows (they're consistent within a (school, day) group in the source).")
    programs = {}  # (school, day) -> dict of program fields
    session_rows = []  # list of (school, day, date_iso, manager_name, manager_mobile)
    for row in sessions_ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        school, region, caterer, date_val, day, mgr, mgr_mob, start, end, dinner, ylevels, building = row
        date_iso = date_val.strftime("%Y-%m-%d") if isinstance(date_val, datetime) else str(date_val)
        key = (school, day)
        if key not in programs:
            yl_list = [int(x.strip()) for x in ylevels.split(",")]
            programs[key] = {
                "school": school,
                "day": day,
                "start_time": parse_time(start),
                "end_time": parse_time(end),
                "dinner_time": parse_time(dinner),
                "building": building,
                "year_levels": yl_list,
                "manager_name": mgr,
                "manager_mobile": mgr_mob,
            }
        session_rows.append((school, day, date_iso, mgr, mgr_mob))

    p("INSERT INTO programs (school_id, day_of_week, start_time, end_time, dinner_time, building, year_levels, manager_name, manager_mobile) VALUES")
    rows = []
    program_id = {}
    pid = 1
    for key, prog in programs.items():
        program_id[key] = pid
        yl_arr = "'{" + ",".join(str(x) for x in prog["year_levels"]) + "}'"
        rows.append(
            f"    ({school_id[prog['school']]}, {sql_str(prog['day'])}, {sql_str(prog['start_time'])}, "
            f"{sql_str(prog['end_time'])}, {sql_str(prog['dinner_time'])}, {sql_str(prog['building'])}, "
            f"{yl_arr}, {sql_str(prog['manager_name'])}, {sql_str(prog['manager_mobile'])})"
        )
        pid += 1
    p(",\n".join(rows) + ";")
    p("")

    # --- sessions (per-date) ---
    p("-- Sessions: one row per dated session from sessions.xlsx. sub_manager_* populated")
    p("-- only when the session's manager differs from the program's regular manager.")
    p("-- Whole-session exclusions (per exclusions.pdf) are skipped here.")
    full_cancellations = {
        (school, date) for school, date, ylevels in EXCLUSIONS if ylevels is None
    }
    p("INSERT INTO sessions (program_id, date, sub_manager_name, sub_manager_mobile) VALUES")
    rows = []
    for school, day, date_iso, mgr, mgr_mob in session_rows:
        if (school, date_iso) in full_cancellations:
            continue
        prog = programs[(school, day)]
        if mgr != prog["manager_name"]:
            sub_name, sub_mob = mgr, mgr_mob
        else:
            sub_name, sub_mob = None, None
        rows.append(
            f"    ({program_id[(school, day)]}, '{date_iso}', {sql_str(sub_name)}, {sql_str(sub_mob)})"
        )
    p(",\n".join(rows) + ";")
    p("")

    # --- students ---
    # Dedupe across sheets by student_email. Track which sheets each student
    # appeared on so we can build the enrolments table next.
    p("-- Students. Deduplicated across all sheets in students.xlsx by student_email.")
    students_wb = openpyxl.load_workbook(RESOURCES / "students.xlsx", data_only=True)
    student_by_email = {}  # email -> dict
    student_sheets = {}    # email -> set of sheet names
    for sheet_name in students_wb.sheetnames:
        ws = students_wb[sheet_name]
        for row in ws.iter_rows(min_row=4, values_only=True):
            name, year, _subjects, dietary_raw, email, parent, p_email, p_mobile = row
            if not name or not email:
                continue
            tags, extra, wants = parse_dietary(dietary_raw)
            if email not in student_by_email:
                student_by_email[email] = {
                    "name": name,
                    "year_level": year,
                    "dietary": tags,
                    "dietary_extra": extra,
                    "wants_catering": wants,
                    "student_email": email,
                    "parent_name": parent,
                    "parent_email": p_email,
                    "parent_mobile": p_mobile,
                }
                student_sheets[email] = set()
            student_sheets[email].add(sheet_name)

    p("INSERT INTO students (name, year_level, dietary, dietary_extra, wants_catering, student_email, parent_name, parent_email, parent_mobile) VALUES")
    rows = []
    student_id = {}
    sid = 1
    for email in student_by_email:
        student_id[email] = sid
        sid += 1
        s = student_by_email[email]
        rows.append(
            f"    ({sql_str(s['name'])}, {s['year_level']}, {sql_array(s['dietary'])}, "
            f"{sql_str(s['dietary_extra'])}, {sql_str(s['wants_catering'])}, "
            f"{sql_str(s['student_email'])}, {sql_str(s['parent_name'])}, "
            f"{sql_str(s['parent_email'])}, {sql_str(s['parent_mobile'])})"
        )
    p(",\n".join(rows) + ";")
    p("")

    # --- enrolments ---
    # Each (student, sheet) maps to one enrolment in the program for that sheet.
    p("-- Enrolments: student → program. A student appearing in multiple sheets is")
    p("-- enrolled in multiple programs (e.g. attends Mon and Tue).")
    p("INSERT INTO enrolments (student_id, program_id) VALUES")
    rows = []
    seen = set()
    for email, sheets in student_sheets.items():
        for sheet_name in sheets:
            school, day = SHEET_TO_PROGRAM[sheet_name]
            key = (student_id[email], program_id[(school, day)])
            if key in seen:
                continue
            seen.add(key)
            rows.append(f"    ({key[0]}, {key[1]})")
    p(",\n".join(rows) + ";")
    p("")

    # Reverse index used by dish-rating generation: which students attend each program.
    program_students = {}  # (school, day) -> set of student emails
    for email, sheets in student_sheets.items():
        for sheet_name in sheets:
            program_students.setdefault(SHEET_TO_PROGRAM[sheet_name], set()).add(email)

    # --- feedback ---
    # Dollar-quoted strings ($$...$$) avoid the need to escape apostrophes in the prose.
    # Shift the whole series forward (FEEDBACK_SHIFT_WEEKS) so it ends in early June 2026;
    # the same shifted dates drive dish_ratings below so feedback and ratings stay aligned.
    feedback_dated = [
        (cname, shift_submitted_at(submitted_at, FEEDBACK_SHIFT_WEEKS), content)
        for cname, submitted_at, content in FEEDBACK
    ]
    p("-- Manager-submitted feedback per caterer. Dish names match the menu so an LLM")
    p("-- can derive per-dish quality signals.")
    p("INSERT INTO feedback (caterer_id, submitted_at, content) VALUES")
    rows = []
    for cname, submitted_at, content in feedback_dated:
        rows.append(f"    ({caterer_id[cname]}, '{submitted_at}', $${content}$$)")
    p(",\n".join(rows) + ";")
    p("")

    # --- dish ratings ---
    # Simulates the per-student 1-10 score a tutor collects during each session and
    # sends back. Each caterer's feedback dates are its past delivery nights; the
    # date's weekday selects which programs ran, so every catering-eligible student
    # in those programs rates one dietary-appropriate dish that night.
    #
    # To keep dishes distinguishable (rather than every average collapsing toward
    # the middle), each dish is given a latent quality drawn from a U-shaped
    # distribution — so dishes lean clearly good or clearly bad — and individual
    # ratings jitter around that quality. A fixed seed keeps the output reproducible.
    random.seed(20260606)

    # Latent per-dish quality on a 1-10 scale. betavariate(0.5, 0.5) is U-shaped,
    # piling mass near the extremes, so dish averages spread across the full range.
    dish_quality = {}
    for cname, items_for_caterer in caterer_items.items():
        cid = caterer_id[cname]
        for name, _tags in items_for_caterer:
            dish_quality[(cid, name)] = 1.0 + 9.0 * random.betavariate(0.5, 0.5)

    rating_rows = []
    seen_ratings = set()  # (student_id, caterer_id, date) — one dish rated per student per night
    for cname, submitted_at, _content in feedback_dated:
        cid = caterer_id[cname]
        date_iso = str(submitted_at).split(" ")[0]
        weekday = datetime.strptime(date_iso, "%Y-%m-%d").strftime("%A")
        items_for_caterer = caterer_items[cname]
        schools_served = [s for s, c in SCHOOL_CATERER.items() if c == cname]
        for school in schools_served:
            for email in sorted(program_students.get((school, weekday), ())):
                s = student_by_email[email]
                if not s["wants_catering"]:
                    continue
                required = set(s["dietary"])
                eligible = [name for name, tags in items_for_caterer if required.issubset(set(tags))]
                if not eligible:
                    continue
                sid_ = student_id[email]
                if (sid_, cid, date_iso) in seen_ratings:
                    continue
                seen_ratings.add((sid_, cid, date_iso))
                dish = random.choice(eligible)
                # Jitter each rating around the dish's quality (tight spread keeps
                # per-dish averages polarised) and clamp into the 1-10 range.
                rating = max(1, min(10, round(random.gauss(dish_quality[(cid, dish)], 1.0))))
                rating_rows.append(f"    ({sid_}, {cid}, {sql_str(dish)}, '{date_iso}', {rating})")

    if rating_rows:
        p("-- Dish ratings: per-student 1-10 scores collected by tutors during each past")
        p("-- session. One dietary-appropriate dish rated per student per delivery night.")
        p("INSERT INTO dish_ratings (student_id, caterer_id, item_name, date, rating) VALUES")
        p(",\n".join(rating_rows) + ";")
    else:
        p("-- (no dish ratings to insert)")
    p("")

    # --- absences ---
    # Absences from absences.pdf plus year-level exclusions converted to per-student absences.
    p("-- Absences: explicit absences from absences.pdf plus year-level exclusions")
    p("-- (exclusions.pdf) expanded into individual absent students.")
    student_id_by_school_name = {}
    for email, sheets in student_sheets.items():
        s = student_by_email[email]
        for sheet_name in sheets:
            school, _ = SHEET_TO_PROGRAM[sheet_name]
            student_id_by_school_name.setdefault((school, s["name"]), []).append(student_id[email])

    absence_rows = []
    seen_abs = set()
    # Explicit absences
    for school, date_iso, sname in ABSENCES:
        ids = student_id_by_school_name.get((school, sname), [])
        if not ids:
            print(f"WARN: absence target {sname!r} not found at {school} on {date_iso}", file=sys.stderr)
            continue
        for sid_ in ids:
            # Find program(s) the student attends at this school
            for (s2, d), pid in program_id.items():
                if s2 != school:
                    continue
                # Only add if there's a session for this program on date_iso
                for srow in session_rows:
                    if srow[0] == school and srow[2] == date_iso and srow[1] == d:
                        key = (sid_, pid, date_iso)
                        if key not in seen_abs:
                            seen_abs.add(key)
                            absence_rows.append(f"    ({sid_}, {pid}, '{date_iso}')")
    # Year-level exclusions
    for school, date_iso, ylevels in EXCLUSIONS:
        if ylevels is None:
            continue
        for srow in session_rows:
            if srow[0] != school or srow[2] != date_iso:
                continue
            day = srow[1]
            pid = program_id[(school, day)]
            for email, sheets in student_sheets.items():
                s = student_by_email[email]
                if s["year_level"] not in ylevels:
                    continue
                # Student must be enrolled in this program
                attends = any(SHEET_TO_PROGRAM[sn] == (school, day) for sn in sheets)
                if not attends:
                    continue
                key = (student_id[email], pid, date_iso)
                if key not in seen_abs:
                    seen_abs.add(key)
                    absence_rows.append(f"    ({student_id[email]}, {pid}, '{date_iso}')")
    if absence_rows:
        p("INSERT INTO absences (student_id, program_id, date) VALUES")
        p(",\n".join(absence_rows) + ";")
    else:
        p("-- (no absences to insert)")
    p("")

    # --- meal orders (student pre-orders) ---
    # A deterministic ~1-in-3 subset of catering-eligible students lock in a specific
    # dish ahead of each upcoming session; everyone else is auto-assigned from the dish
    # ranking at order time. Eligibility mirrors dish_ratings: enrolled in the program,
    # wants_catering, not absent for that session, and a dietary-appropriate dish exists
    # on the school's caterer menu. The chosen dish is FK-safe (it's on that caterer).
    order_rows = []
    seen_orders = set()  # (student_id, program_id, date) — one pre-order per student per session
    for school, day, date_iso, _mgr, _mgr_mob in session_rows:
        if (school, date_iso) in full_cancellations:
            continue
        pid = program_id[(school, day)]
        cname = SCHOOL_CATERER[school]
        cid = caterer_id[cname]
        items_for_caterer = caterer_items[cname]
        for email in sorted(program_students.get((school, day), ())):
            s = student_by_email[email]
            if not s["wants_catering"]:
                continue
            sid_ = student_id[email]
            if (sid_, pid, date_iso) in seen_abs:          # absent that session
                continue
            if (sid_, pid, date_iso) in seen_orders:
                continue
            required = set(s["dietary"])
            eligible = [name for name, tags in items_for_caterer if required.issubset(set(tags))]
            if not eligible:
                continue
            if random.random() >= 0.34:                    # ~1 in 3 pre-orders
                continue
            seen_orders.add((sid_, pid, date_iso))
            dish = random.choice(eligible)
            order_rows.append(f"    ({sid_}, {pid}, '{date_iso}', {cid}, {sql_str(dish)})")

    if order_rows:
        p("-- Student pre-ordered meals: a locked-in dish for an upcoming session, for a")
        p("-- subset of eligible students; the rest are auto-assigned from the dish ranking.")
        p("INSERT INTO meal_orders (student_id, program_id, date, caterer_id, item_name) VALUES")
        p(",\n".join(order_rows) + ";")
    else:
        p("-- (no meal orders to insert)")
    p("")

    return "\n".join(out)


if __name__ == "__main__":
    output_path = Path(__file__).resolve().parent / "populate2.sql"
    output_path.write_text(main(), encoding="utf-8")
    print(f"Wrote {output_path}", file=sys.stderr)
